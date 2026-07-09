"""v3.3 — Project Insight, Section 9: Executive Forecast Reports.

Reuses the exact export pattern `atlas_report_service.py` (Atlas v3.1)
established: CSV (`csv.DictWriter`/`StringIO`), XLSX
(`openpyxl.Workbook`/`BytesIO`), PDF (`reportlab.pdfgen.canvas.Canvas`/
`BytesIO`) — all in-memory, no temp files, no new export library added.
"""
from __future__ import annotations

import csv
import json
import uuid
from datetime import datetime, timezone
from io import BytesIO, StringIO

from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from sqlalchemy.orm import Session

from app.models.predictive_insight import DISCLAIMER, REPORT_CADENCES, ExecutiveForecastReport
from app.services import (
    insight_instrument_forecast_service,
    insight_operational_forecast_service,
    insight_quality_trend_service,
    insight_recommendation_service,
)

_CADENCE_TO_HORIZON = {"weekly": "7_day", "monthly": "30_day", "quarterly": "90_day", "annual": "rolling_annual"}


def _row_to_dict(obj) -> dict:
    result: dict = {}
    for col in obj.__table__.columns:
        val = getattr(obj, col.name)
        if hasattr(val, "isoformat"):
            val = val.isoformat()
        result[col.name] = val
    return result


def _period_label(cadence: str) -> str:
    now = datetime.now(timezone.utc)
    if cadence == "annual":
        return f"{now.year}"
    if cadence == "quarterly":
        return f"{now.year}-Q{(now.month - 1) // 3 + 1}"
    if cadence == "weekly":
        iso = now.isocalendar()
        return f"{iso[0]}-W{iso[1]:02d}"
    return f"{now.year}-{now.month:02d}"


def _build_summary(db: Session, tenant_id: str, cadence: str) -> dict:
    horizon = _CADENCE_TO_HORIZON[cadence]
    quality_trends = insight_quality_trend_service.generate_all_quality_trend_forecasts(db, tenant_id, horizon=horizon)
    operational = insight_operational_forecast_service.generate_all_operational_forecasts(db, tenant_id, horizon=horizon)
    instrument_lifecycle = insight_instrument_forecast_service.generate_lifecycle_forecasts_for_tenant(db, tenant_id)
    recommendations = insight_recommendation_service.generate_recommendations(db, tenant_id)

    return {
        "cadence": cadence, "horizon": horizon,
        "quality_trend_forecasts": [
            {"metric": t["metric"], "trend_direction": t["trend_direction"], "forecast_value": t["forecast_value"], "confidence_level": t["confidence_level"]}
            for t in quality_trends
        ],
        "operational_forecasts": [
            {"forecast_type": o["forecast_type"], "forecast_value": o["forecast_value"], "confidence_level": o["confidence_level"]}
            for o in operational
        ],
        "instrument_lifecycle_forecasts": [
            {"instrument_type": f["instrument_type"], "lifecycle_risk_tier": f["lifecycle_risk_tier"], "retirement_likelihood": f["retirement_likelihood"]}
            for f in instrument_lifecycle
        ],
        "open_recommendations": [r for r in recommendations if r["status"] == "open"],
    }


def generate_executive_forecast_report(db: Session, tenant_id: str, *, cadence: str, generated_by: str = "system") -> dict:
    if cadence not in REPORT_CADENCES:
        raise ValueError(f"cadence must be one of {REPORT_CADENCES}")

    summary = _build_summary(db, tenant_id, cadence)
    period_label = _period_label(cadence)

    row = ExecutiveForecastReport(
        tenant_id=tenant_id, report_ref=f"IFR-{datetime.now(timezone.utc).year}-{uuid.uuid4().hex[:6].upper()}",
        cadence=cadence, period_label=period_label,
        title=f"{cadence.title()} Forecast Report — {period_label}",
        summary_json=json.dumps(summary, default=str), generated_by=generated_by,
    )
    db.add(row)
    db.commit()
    db.refresh(row)

    result = _row_to_dict(row)
    result["summary"] = summary
    return result


def get_report(db: Session, tenant_id: str, report_id: int) -> dict | None:
    row = db.query(ExecutiveForecastReport).filter(ExecutiveForecastReport.id == report_id, ExecutiveForecastReport.tenant_id == tenant_id).first()
    if row is None:
        return None
    result = _row_to_dict(row)
    result["summary"] = json.loads(row.summary_json)
    return result


def list_reports(db: Session, tenant_id: str, *, cadence: str = "") -> list[dict]:
    q = db.query(ExecutiveForecastReport).filter(ExecutiveForecastReport.tenant_id == tenant_id)
    if cadence:
        q = q.filter(ExecutiveForecastReport.cadence == cadence)
    rows = q.order_by(ExecutiveForecastReport.id.desc()).all()
    return [_row_to_dict(r) for r in rows]


def build_report_csv_bytes(report: dict) -> bytes:
    output = StringIO()
    rows = report["summary"].get("quality_trend_forecasts") or []
    if rows:
        writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    return output.getvalue().encode("utf-8")


def build_report_xlsx_bytes(report: dict) -> bytes:
    wb = Workbook()
    summary = report["summary"]

    ws = wb.active
    ws.title = "Quality Trends"
    quality_rows = summary.get("quality_trend_forecasts") or []
    if quality_rows:
        headers = list(quality_rows[0].keys())
        ws.append(headers)
        for item in quality_rows:
            ws.append([item.get(h, "") for h in headers])

    ws2 = wb.create_sheet("Operational Forecasts")
    op_rows = summary.get("operational_forecasts") or []
    if op_rows:
        headers = list(op_rows[0].keys())
        ws2.append(headers)
        for item in op_rows:
            ws2.append([item.get(h, "") for h in headers])

    ws3 = wb.create_sheet("Instrument Lifecycle")
    lifecycle_rows = summary.get("instrument_lifecycle_forecasts") or []
    if lifecycle_rows:
        headers = list(lifecycle_rows[0].keys())
        ws3.append(headers)
        for item in lifecycle_rows:
            ws3.append([item.get(h, "") for h in headers])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def build_report_pdf_bytes(report: dict) -> bytes:
    bio = BytesIO()
    c = canvas.Canvas(bio, pagesize=letter)
    width, height = letter
    y = height - 50

    c.setFont("Helvetica-Bold", 18)
    c.drawString(50, y, report["title"])
    y -= 26
    c.setFont("Helvetica", 10)
    c.drawString(50, y, f"Period: {report['period_label']} · Cadence: {report['cadence']}")
    y -= 24

    summary = report["summary"]
    c.setFont("Helvetica-Bold", 12)
    c.drawString(50, y, "Quality Trend Forecasts")
    y -= 18
    c.setFont("Helvetica", 10)
    for t in summary.get("quality_trend_forecasts", [])[:12]:
        c.drawString(60, y, f"{t['metric']}: {t['trend_direction']} (forecast {t['forecast_value']}, confidence {t['confidence_level']})")
        y -= 14

    y -= 10
    c.setFont("Helvetica-Oblique", 8)
    for line in [DISCLAIMER[i:i + 100] for i in range(0, len(DISCLAIMER), 100)]:
        c.drawString(50, y, line)
        y -= 10

    c.showPage()
    c.save()
    bio.seek(0)
    return bio.getvalue()
