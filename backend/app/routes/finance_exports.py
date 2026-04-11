from __future__ import annotations

from io import BytesIO, StringIO
import csv
import json
import zipfile

from fastapi import APIRouter, Depends
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook

from app.billing import build_invoice_preview
from app.deps import get_db
from app.db import models
from app.tenant import resolve_tenant
from app.tenant_authz import require_tenant_roles

router = APIRouter(tags=["finance-exports"])


def _csv_text(items: list[dict]) -> str:
    output = StringIO()
    if items:
        writer = csv.DictWriter(output, fieldnames=list(items[0].keys()))
        writer.writeheader()
        writer.writerows(items)
    return output.getvalue()


def _xlsx_bytes(preview: dict, payments: list[dict], invoices: list[dict]) -> bytes:
    wb = Workbook()

    ws = wb.active
    ws.title = "Invoice Preview"
    ws.append(["field", "value"])
    for key, value in preview.items():
        if isinstance(value, (dict, list)):
            ws.append([key, json.dumps(value)])
        else:
            ws.append([key, value])

    ws2 = wb.create_sheet("Invoice History")
    if invoices:
        headers = list(invoices[0].keys())
        ws2.append(headers)
        for item in invoices:
            ws2.append([item.get(h, "") for h in headers])

    ws3 = wb.create_sheet("Payment History")
    if payments:
        headers = list(payments[0].keys())
        ws3.append(headers)
        for item in payments:
            ws3.append([item.get(h, "") for h in headers])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


@router.get("/finance-console/invoice-preview.json")
def finance_invoice_preview_json(
    tenant: dict = Depends(resolve_tenant),
    db: Session = Depends(get_db),
    current_user=Depends(require_tenant_roles("tenant_admin", "site_admin")),
):
    return JSONResponse(build_invoice_preview(db, tenant["tenant_id"], tenant["tenant_name"]))


@router.get("/finance-console/invoice-preview.csv")
def finance_invoice_preview_csv(
    tenant: dict = Depends(resolve_tenant),
    db: Session = Depends(get_db),
    current_user=Depends(require_tenant_roles("tenant_admin", "site_admin")),
):
    preview = build_invoice_preview(db, tenant["tenant_id"], tenant["tenant_name"])
    items = preview["line_items"]
    return StreamingResponse(
        iter([_csv_text(items)]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=lumenai_{tenant['tenant_id']}_invoice_preview.csv"},
    )


@router.get("/finance-console/invoice-preview.bundle.zip")
def finance_invoice_preview_bundle(
    tenant: dict = Depends(resolve_tenant),
    db: Session = Depends(get_db),
    current_user=Depends(require_tenant_roles("tenant_admin", "site_admin")),
):
    preview = build_invoice_preview(db, tenant["tenant_id"], tenant["tenant_name"])

    invoice_rows = (
        db.query(models.InvoiceLineItem)
        .filter(models.InvoiceLineItem.tenant_id == tenant["tenant_id"])
        .order_by(models.InvoiceLineItem.id.desc())
        .limit(100)
        .all()
    )
    invoices = [
        {
            "id": row.id,
            "billing_month": row.billing_month,
            "item_type": row.item_type,
            "quantity": row.quantity,
            "unit_price_cents": row.unit_price_cents,
            "amount_cents": row.amount_cents,
            "notes": row.notes,
            "created_at": row.created_at.isoformat() if row.created_at else None,
        }
        for row in invoice_rows
    ]

    payment_rows = (
        db.query(models.PaymentEvent)
        .filter(models.PaymentEvent.tenant_id == tenant["tenant_id"])
        .order_by(models.PaymentEvent.id.desc())
        .limit(50)
        .all()
    )
    payments = [
        {
            "id": row.id,
            "event_type": row.event_type,
            "status": row.status,
            "amount_cents": row.amount_cents,
            "billing_month": row.billing_month,
            "notes": row.notes,
            "created_at": row.created_at.isoformat() if row.created_at else None,
        }
        for row in payment_rows
    ]

    xlsx = _xlsx_bytes(preview, payments, invoices)
    bio = BytesIO()
    with zipfile.ZipFile(bio, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"lumenai_{tenant['tenant_id']}_invoice_preview.json", json.dumps(preview, indent=2))
        zf.writestr(f"lumenai_{tenant['tenant_id']}_invoice_preview.csv", _csv_text(preview["line_items"]))
        zf.writestr(f"lumenai_{tenant['tenant_id']}_payments.json", json.dumps(payments, indent=2))
        zf.writestr(f"lumenai_{tenant['tenant_id']}_invoices.json", json.dumps(invoices, indent=2))
        zf.writestr(f"lumenai_{tenant['tenant_id']}_finance_console.xlsx", xlsx)

    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=lumenai_{tenant['tenant_id']}_finance_console_bundle.zip"},
    )
