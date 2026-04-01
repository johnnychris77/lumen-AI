from __future__ import annotations

from io import BytesIO, StringIO
from datetime import datetime, timedelta, timezone
import csv
import json
import zipfile

from fastapi import APIRouter, Depends, Query
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.compliance_signing import sign_manifest, verify_manifest
from app.deps import get_db
from app.db import models
from app.tenant import resolve_tenant
from app.tenant_authz import require_tenant_roles

router = APIRouter(tags=["compliance-exports"])


def _within_days(dt: datetime | None, days: int) -> bool:
    if not dt:
        return False
    now = datetime.now(timezone.utc)
    return dt >= now - timedelta(days=days)


def _tenant_inspections(db: Session, tenant_id: str, days: int):
    rows = (
        db.query(models.Inspection)
        .filter(models.Inspection.tenant_id == tenant_id)
        .order_by(models.Inspection.id.desc())
        .all()
    )
    return [r for r in rows if _within_days(r.created_at, days)]


def _tenant_audit_logs(db: Session, tenant_id: str, days: int):
    rows = (
        db.query(models.AuditLog)
        .filter(models.AuditLog.tenant_id == tenant_id)
        .order_by(models.AuditLog.id.desc())
        .all()
    )
    return [r for r in rows if _within_days(r.created_at, days)]


def _inspection_items(rows: list[models.Inspection]) -> list[dict]:
    return [
        {
            "inspection_id": r.id,
            "tenant_id": r.tenant_id,
            "tenant_name": r.tenant_name,
            "site_name": r.site_name,
            "vendor_name": r.vendor_name,
            "file_name": r.file_name,
            "status": r.status,
            "risk_score": r.risk_score,
            "detected_issue": r.detected_issue,
            "alert_status": getattr(r, "alert_status", ""),
            "qa_review_status": getattr(r, "qa_review_status", ""),
            "created_at": r.created_at.isoformat() if r.created_at else None,
        }
        for r in rows
    ]


def _audit_items(rows: list[models.AuditLog]) -> list[dict]:
    return [
        {
            "audit_id": r.id,
            "tenant_id": r.tenant_id,
            "tenant_name": r.tenant_name,
            "actor_email": r.actor_email,
            "actor_role": r.actor_role,
            "action_type": r.action_type,
            "resource_type": r.resource_type,
            "resource_id": r.resource_id,
            "status": r.status,
            "request_method": r.request_method,
            "request_path": r.request_path,
            "client_ip": r.client_ip,
            "details": r.details,
            "compliance_flag": r.compliance_flag,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        }
        for r in rows
    ]


def _summary(tenant: dict, inspections: list[dict], audit_logs: list[dict]) -> dict:
    total_inspections = len(inspections)
    total_audit_events = len(audit_logs)
    open_alerts = sum(1 for x in inspections if (x.get("alert_status") or "").lower() != "resolved")
    high_risk = sum(1 for x in inspections if int(x.get("risk_score") or 0) >= 80)
    compliance_events = sum(1 for x in audit_logs if bool(x.get("compliance_flag")))

    return {
        "tenant_id": tenant["tenant_id"],
        "tenant_name": tenant["tenant_name"],
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "total_inspections": total_inspections,
        "total_audit_events": total_audit_events,
        "open_alerts": open_alerts,
        "high_risk_count": high_risk,
        "compliance_flagged_events": compliance_events,
    }


def _csv_text(items: list[dict]) -> str:
    output = StringIO()
    if items:
        writer = csv.DictWriter(output, fieldnames=list(items[0].keys()))
        writer.writeheader()
        writer.writerows(items)
    return output.getvalue()


def _xlsx_bytes(summary: dict, inspections: list[dict], audit_logs: list[dict], manifest: dict) -> bytes:
    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    ws.append(["metric", "value"])
    for k, v in summary.items():
        ws.append([k, v])

    ws_manifest = wb.create_sheet("Manifest")
    ws_manifest.append(["field", "value"])
    for k, v in manifest.items():
        if k == "integrity":
            for ik, iv in v.items():
                ws_manifest.append([f"integrity.{ik}", iv])
        else:
            ws_manifest.append([k, json.dumps(v) if isinstance(v, (dict, list)) else v])

    ws2 = wb.create_sheet("Inspections")
    if inspections:
        headers = list(inspections[0].keys())
        ws2.append(headers)
        for item in inspections:
            ws2.append([item.get(h, "") for h in headers])

    ws3 = wb.create_sheet("Audit Logs")
    if audit_logs:
        headers = list(audit_logs[0].keys())
        ws3.append(headers)
        for item in audit_logs:
            ws3.append([item.get(h, "") for h in headers])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def _manifest(summary: dict, inspections: list[dict], audit_logs: list[dict]) -> dict:
    payload = {
        "summary": summary,
        "inspection_count": len(inspections),
        "audit_log_count": len(audit_logs),
    }
    integrity = sign_manifest(payload)
    return {
        **payload,
        "integrity": integrity,
        "immutability_note": "This evidence pack is export-time immutable. Verify integrity with the included hash and signature.",
    }


@router.get("/compliance-exports/evidence-pack.json")
def compliance_evidence_pack_json(
    days: int = Query(default=30, ge=1, le=365),
    tenant: dict = Depends(resolve_tenant),
    db: Session = Depends(get_db),
    current_user=Depends(require_tenant_roles("tenant_admin", "site_admin")),
):
    inspections = _inspection_items(_tenant_inspections(db, tenant["tenant_id"], days))
    audit_logs = _audit_items(_tenant_audit_logs(db, tenant["tenant_id"], days))
    summary = _summary(tenant, inspections, audit_logs)
    manifest = _manifest(summary, inspections, audit_logs)

    return JSONResponse({
        "manifest": manifest,
        "summary": summary,
        "inspections": inspections,
        "audit_logs": audit_logs,
    })


@router.get("/compliance-exports/evidence-pack.csv")
def compliance_evidence_pack_csv(
    days: int = Query(default=30, ge=1, le=365),
    tenant: dict = Depends(resolve_tenant),
    db: Session = Depends(get_db),
    current_user=Depends(require_tenant_roles("tenant_admin", "site_admin")),
):
    inspections = _inspection_items(_tenant_inspections(db, tenant["tenant_id"], days))
    text = _csv_text(inspections)
    return StreamingResponse(
        iter([text]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=lumenai_{tenant['tenant_id']}_evidence_pack_inspections.csv"},
    )


@router.get("/compliance-exports/evidence-pack.xlsx")
def compliance_evidence_pack_xlsx(
    days: int = Query(default=30, ge=1, le=365),
    tenant: dict = Depends(resolve_tenant),
    db: Session = Depends(get_db),
    current_user=Depends(require_tenant_roles("tenant_admin", "site_admin")),
):
    inspections = _inspection_items(_tenant_inspections(db, tenant["tenant_id"], days))
    audit_logs = _audit_items(_tenant_audit_logs(db, tenant["tenant_id"], days))
    summary = _summary(tenant, inspections, audit_logs)
    manifest = _manifest(summary, inspections, audit_logs)
    content = _xlsx_bytes(summary, inspections, audit_logs, manifest)
    return StreamingResponse(
        iter([content]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=lumenai_{tenant['tenant_id']}_evidence_pack.xlsx"},
    )


@router.get("/compliance-exports/evidence-pack.bundle.zip")
def compliance_evidence_pack_bundle(
    days: int = Query(default=30, ge=1, le=365),
    tenant: dict = Depends(resolve_tenant),
    db: Session = Depends(get_db),
    current_user=Depends(require_tenant_roles("tenant_admin", "site_admin")),
):
    inspections = _inspection_items(_tenant_inspections(db, tenant["tenant_id"], days))
    audit_logs = _audit_items(_tenant_audit_logs(db, tenant["tenant_id"], days))
    summary = _summary(tenant, inspections, audit_logs)
    manifest = _manifest(summary, inspections, audit_logs)

    json_payload = {
        "manifest": manifest,
        "summary": summary,
        "inspections": inspections,
        "audit_logs": audit_logs,
    }

    xlsx_content = _xlsx_bytes(summary, inspections, audit_logs, manifest)
    inspections_csv = _csv_text(inspections)
    audits_csv = _csv_text(audit_logs)
    manifest_json = json.dumps(manifest, indent=2)
    payload_json = json.dumps(json_payload, indent=2)

    bio = BytesIO()
    with zipfile.ZipFile(bio, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"lumenai_{tenant['tenant_id']}_manifest.json", manifest_json)
        zf.writestr(f"lumenai_{tenant['tenant_id']}_evidence_pack.json", payload_json)
        zf.writestr(f"lumenai_{tenant['tenant_id']}_inspections.csv", inspections_csv)
        zf.writestr(f"lumenai_{tenant['tenant_id']}_audit_logs.csv", audits_csv)
        zf.writestr(f"lumenai_{tenant['tenant_id']}_evidence_pack.xlsx", xlsx_content)

    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=lumenai_{tenant['tenant_id']}_evidence_pack_bundle.zip"},
    )


@router.post("/compliance-exports/verify")
def verify_evidence_pack(
    payload: dict,
    tenant: dict = Depends(resolve_tenant),
    current_user=Depends(require_tenant_roles("tenant_admin", "site_admin")),
):
    manifest = payload.get("manifest", {})
    summary = payload.get("summary", {})
    inspections = payload.get("inspections", [])
    audit_logs = payload.get("audit_logs", [])

    verification_payload = {
        "summary": {
            "tenant_id": summary.get("tenant_id"),
            "tenant_name": summary.get("tenant_name"),
            "generated_at": summary.get("generated_at"),
            "total_inspections": summary.get("total_inspections"),
            "total_audit_events": summary.get("total_audit_events"),
            "open_alerts": summary.get("open_alerts"),
            "high_risk_count": summary.get("high_risk_count"),
            "compliance_flagged_events": summary.get("compliance_flagged_events"),
        },
        "inspection_count": len(inspections),
        "audit_log_count": len(audit_logs),
    }

    verification = verify_manifest(verification_payload, {"integrity": manifest.get("integrity", {})})
    return JSONResponse({
        "tenant_id": tenant["tenant_id"],
        "tenant_name": tenant["tenant_name"],
        "verification": verification,
    })
